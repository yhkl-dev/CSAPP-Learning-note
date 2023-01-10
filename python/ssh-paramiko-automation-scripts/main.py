import sys
from collections import namedtuple
from typing import Any, Dict, List
import time
import openpyxl
import paramiko

from openpyxl import cell


IP = namedtuple("IP", ["ip", "port", "username", "password", "cmd"])
Command = namedtuple("Command", ["command", "description"])


def init_ssh_client(ip: IP) -> paramiko.SSHClient:
    ssh_client = paramiko.SSHClient()
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    ssh_client.connect(
        hostname=ip.ip, port=ip.port, username=ip.username, password=ip.password
    )
    return ssh_client


def exec_command_with_invoke_shell(ssh_client: paramiko.SSHClient, command: str) -> str:
    session = ssh_client.invoke_shell()
    session.send(f"{str(command).strip()}" + "\n")
    time.sleep(2)
    while not session.recv_ready():
        time.sleep(0.02)
    data = session.recv(65535).decode()
    return cell.cell.ILLEGAL_CHARACTERS_RE.sub(r" ", data)


def exec_command(ssh_client: paramiko.SSHClient, command: str) -> str:
    _, stdout, stderr = ssh_client.exec_command(command)
    stdout, stderr = stdout.read(), stderr.read()
    results = stdout if stdout else stderr
    return results.decode("utf-8")


def read_ips(filename: str) -> List[IP]:
    book = openpyxl.load_workbook(filename, read_only=True)
    sheet = book.active
    ips = []
    for x in range(2, sheet.max_row + 1):
        cmd = sheet["E" + str(x)].value
        if not cmd.endswith(".xlsx"):
            cmd += ".xlsx"
        item = [
            str(sheet["A" + str(x)].value),
            int(sheet["B" + str(x)].value),
            str(sheet["C" + str(x)].value),
            str(sheet["D" + str(x)].value),
            cmd,
        ]
        ips.append(IP(*item))
    return ips


def read_commands(filename: str) -> List[str]:
    book = openpyxl.load_workbook(filename, read_only=True)
    sheet = book.active
    commands = []
    for x in range(2, sheet.max_row + 1):
        item = [
            str(sheet["A" + str(x)].value),
            str(sheet["B" + str(x)].value),
        ]
        commands.append(Command(*item))
    return commands


def dispatch_commands(ips: List[IP]) -> List[Dict[str, str]]:
    results = []
    for ip in ips:
        commands = read_commands(ip.cmd)
        ssh_client = init_ssh_client(ip)
        command_result = {
            "result_file": f"output_results_{ip.cmd.split('.xlsx')[0]}.xlsx",
            "ip": ip.ip,
        }
        for command in commands:
            command_result[command.description] = exec_command_with_invoke_shell(
                ssh_client, command.command
            )
        results.append(command_result)
        ssh_client.close()
    return results


def write_results(results: List[Dict[str, Any]]) -> None:
    def deal_results(results) -> Dict[str, List[Dict[str, str]]]:
        r = dict()
        for result in results:
            r[result["result_file"]] = r.get(result["result_file"], [])
            r[result["result_file"]].append(
                {key: value for key, value in result.items() if key != "result_file"}
            )

        return r

    res = deal_results(results)
    for file_name, result in res.items():
        workbook = openpyxl.Workbook()
        default = workbook["Sheet"]
        workbook.remove(default)
        worksheet = workbook.create_sheet("sheet1")
        worksheet.append(list(result[0].keys()))
        for r in result:
            worksheet.append(list(r.values()))

        workbook.save(file_name)
        print(f"result saved successfully in file: {file_name}")


def main(filename: str) -> None:
    ips = read_ips(filename)
    results = dispatch_commands(ips)
    write_results(results)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("filename must be specified")
        print("")
        print("Usage:")
        print("     python main.py <filename>")
        sys.exit(1)

    filename = sys.argv[1]
    if not filename.endswith(".xlsx"):
        print("file must be xlsx file type")
        sys.exit(1)

    main(filename)
